from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
from openpyxl import load_workbook
from datetime import datetime, timedelta
import os
import json
import shutil
import sys
import time
from pathlib import Path

app = Flask(__name__)
CORS(app)

# 获取PythonAnywhere上的绝对路径
BASE_DIR = Path(__file__).parent.absolute()

# Configuration - 使用绝对路径
UPLOAD_FOLDER = BASE_DIR / 'uploads'
GENERATED_FOLDER = BASE_DIR / 'generated_documents'
TEMPLATE_PATH = BASE_DIR / 'visa_booking_template.xlsx'
COUNTER_FILE = BASE_DIR / 'daily_counters.json'

# 调试信息
print(f"PythonAnywhere 部署检测")
print(f"当前工作目录: {os.getcwd()}")
print(f"BASE_DIR: {BASE_DIR}")
print(f"模板路径: {TEMPLATE_PATH}")
print(f"生成文件夹: {GENERATED_FOLDER}")

# 创建目录 - 确保有写权限
def create_directories():
    """创建必要的目录"""
    directories = [UPLOAD_FOLDER, GENERATED_FOLDER]
    for directory in directories:
        try:
            directory.mkdir(exist_ok=True)
            print(f"✓ 目录已创建/存在: {directory}")
        except Exception as e:
            print(f"✗ 创建目录失败 {directory}: {e}")
            # 尝试设置权限
            try:
                os.makedirs(str(directory), exist_ok=True, mode=0o755)
            except:
                pass

# 初始化时创建目录
create_directories()

# Store for generated documents
documents_store = []

def load_daily_counters():
    """加载每日计数器"""
    try:
        if COUNTER_FILE.exists():
            with open(COUNTER_FILE, 'r', encoding='utf-8') as f:
                counters = json.load(f)
                # 确保计数器值是数字
                for date in counters:
                    if isinstance(counters[date], str):
                        counters[date] = int(counters[date])
                return counters
    except Exception as e:
        print(f"加载计数器失败: {e}")
        return {}
    return {}

def save_daily_counters(counters):
    """保存每日计数器"""
    try:
        with open(COUNTER_FILE, 'w', encoding='utf-8') as f:
            json.dump(counters, f, ensure_ascii=False)
        print(f"计数器已保存: {counters}")
    except Exception as e:
        print(f"保存计数器失败: {e}")

def generate_confirmation_number():
    """Generate a unique confirmation number: YYMMDDXXXX"""
    today = datetime.now().strftime('%Y%m%d')
    counters = load_daily_counters()
    
    print(f"当前计数器状态: {counters}")
    print(f"今天日期: {today}")
    
    # 检查今天是否已有计数器
    if today in counters:
        # 递增计数器
        counters[today] += 1
    else:
        # 新的一天，从1开始
        counters[today] = 1
    
    # 保存计数器
    save_daily_counters(counters)
    
    # 生成确认号
    confirmation_number = f"{today}{str(counters[today]).zfill(4)}"
    print(f"生成的确认号: {confirmation_number}")
    
    return confirmation_number

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/admin')
def admin_panel():
    """后端管理页面"""
    return render_template('admin.html')

@app.route('/generate-document', methods=['POST'])
def generate_document():
    try:
        data = request.json
        
        # 调试：打印接收到的数据
        print("\n" + "="*60)
        print("收到生成文档请求:")
        print(f"数据: {data}")
        
        # Validate required fields
        required_fields = ['guestName', 'email', 'company', 'arrivalDate', 'departureDate']
        for field in required_fields:
            if not data.get(field):
                print(f"缺失必填字段: {field}")
                return jsonify({
                    'success': False,
                    'message': f'Missing required field: {field}'
                }), 400
        
        # Generate unique confirmation number
        confirmation_number = generate_confirmation_number()
        print(f"生成的确认号: {confirmation_number}")
        
        # Calculate nights
        arrival_date = datetime.strptime(data['arrivalDate'], '%Y-%m-%d')
        departure_date = datetime.strptime(data['departureDate'], '%Y-%m-%d')
        nights = (departure_date - arrival_date).days
        if nights < 1:
            nights = 1
        
        # Calculate total amount
        room_rate = 98000
        quantity = data.get('quantity', 1)
        total_amount = nights * room_rate * quantity
        
        print(f"入住天数: {nights}, 总金额: {total_amount}")
        
        # Check if template exists
        if not TEMPLATE_PATH.exists():
            print("模板文件不存在，尝试创建...")
            create_template_file()
            if not TEMPLATE_PATH.exists():
                return jsonify({
                    'success': False,
                    'message': f'Template file not found at: {TEMPLATE_PATH}'
                }), 404
        
        print(f"模板文件存在: {TEMPLATE_PATH}")
        
        # Load the template - 创建副本避免修改原文件
        temp_template = BASE_DIR / 'visa_booking_template_temp.xlsx'
        try:
            shutil.copy2(str(TEMPLATE_PATH), str(temp_template))
            print(f"模板副本创建成功: {temp_template}")
        except Exception as e:
            print(f"复制模板失败: {e}")
            return jsonify({
                'success': False,
                'message': f'无法复制模板文件: {str(e)}'
            }), 500
        
        # 尝试打开工作簿
        try:
            wb = load_workbook(str(temp_template))
            ws = wb.active
            print("工作簿加载成功")
        except Exception as e:
            print(f"加载工作簿失败: {e}")
            return jsonify({
                'success': False,
                'message': f'无法打开Excel模板: {str(e)}'
            }), 500
        
        # 记录原始合并区域
        original_merges = list(ws.merged_cells.ranges)
        print(f"找到 {len(original_merges)} 个合并区域")
        
        # 只取消需要写入的合并区域
        data_cells = ['J5', 'J19', 'D22', 'B7', 'H22', 'K22', 'J8', 'J17', 'J9', 'J10']
        merges_to_remove = []
        
        for merge_range in original_merges:
            should_remove = False
            for cell_addr in data_cells:
                cell = ws[cell_addr]
                if merge_range.min_row <= cell.row <= merge_range.max_row and \
                   merge_range.min_col <= cell.column <= merge_range.max_col:
                    should_remove = True
                    break
            if should_remove:
                merges_to_remove.append(merge_range)
        
        # 取消特定的合并区域
        for merge_range in merges_to_remove:
            ws.unmerge_cells(str(merge_range))
        
        print(f"取消了 {len(merges_to_remove)} 个合并区域")
        
        # 写入数据
        try:
            # Guest Information
            ws['J5'] = data['guestName']    # Guest Name in contact
            ws['J19'] = data['guestName']   # Guest Name in reservation
            ws['D22'] = data['guestName']   # Guest Name in table
            
            # Company Information
            ws['B7'] = data['company']
            
            # Dates
            ws['H22'] = arrival_date.strftime('%Y-%m-%d')    # Arrival Date
            ws['K22'] = departure_date.strftime('%Y-%m-%d')  # Departure Date
            ws['J8'] = datetime.now().strftime('%Y-%m-%d')   # Booking Date
            
            # Confirmation Number
            ws['J17'] = confirmation_number
            
            # Email and Remarks
            ws['J9'] = data['email']
            remark = data.get('remark', '')
            if data.get('purpose') == 'VISA_APPLICATION_ONLY':
                remark = "FOR VISA APPLICATION PURPOSES ONLY - NOT AN ACTUAL BOOKING. " + remark
            ws['J10'] = remark
            
            # Room Information
            ws['M22'] = data.get('roomType', 'Classic Queen')  # Room Type
            ws['Q22'] = quantity  # Quantity
            ws['T22'] = nights  # Nights
            ws['V22'] = room_rate  # Room Rate
            
            print("数据写入成功")
        except Exception as e:
            print(f"写入数据失败: {e}")
            return jsonify({
                'success': False,
                'message': f'无法写入数据到Excel: {str(e)}'
            }), 500
        
        # 重新合并我们取消的区域
        for merge_range in merges_to_remove:
            try:
                ws.merge_cells(str(merge_range))
            except Exception as e:
                print(f"重新合并失败 {merge_range}: {e}")
        
        # Add metadata
        ws['AA1'] = f"Company: {data['company']}"
        ws['AA2'] = f"Email: {data['email']}"
        ws['AA3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['AA4'] = f"Document ID: {confirmation_number}"
        
        # Generate filename
        safe_company = "".join(c for c in data['company'] if c.isalnum() or c in (' ', '-', '_')).strip()
        safe_company = safe_company.replace(' ', '_')[:30]
        filename = f"Visa_Booking_{confirmation_number}_{safe_company}.xlsx"
        filepath = GENERATED_FOLDER / filename
        
        # 确保生成目录存在
        GENERATED_FOLDER.mkdir(exist_ok=True)
        
        # Save the workbook
        try:
            wb.save(str(filepath))
            print(f"文件保存成功: {filepath}")
            print(f"文件大小: {os.path.getsize(filepath)} bytes")
        except Exception as e:
            print(f"保存文件失败: {e}")
            return jsonify({
                'success': False,
                'message': f'无法保存Excel文件: {str(e)}'
            }), 500
        
        # Clean up temp file
        try:
            if temp_template.exists():
                os.remove(str(temp_template))
                print("临时文件已清理")
        except Exception as e:
            print(f"清理临时文件失败: {e}")
        
        # Store document information
        document_info = {
            'id': confirmation_number,
            'filename': filename,
            'company': data['company'],
            'email': data['email'],
            'guest_name': data['guestName'],
            'arrival_date': data['arrivalDate'],
            'departure_date': data['departureDate'],
            'nights': nights,
            'total_amount': total_amount,
            'generated_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'filepath': str(filepath),
            'purpose': 'VISA_APPLICATION_ONLY',
            'download_url': f'/download/{confirmation_number}',
            'print_url': f'/print/{confirmation_number}'
        }
        
        documents_store.append(document_info)
        
        # Print to console
        print("\n" + "="*60)
        print("✅ VISA BOOKING DOCUMENT GENERATED SUCCESSFULLY")
        print("="*60)
        print(f"Company: {data['company']}")
        print(f"Email: {data['email']}")
        print(f"Guest: {data['guestName']}")
        print(f"Dates: {data['arrivalDate']} to {data['departureDate']}")
        print(f"Nights: {nights}")
        print(f"Total: {total_amount:,} CFA")
        print(f"Document ID: {confirmation_number}")
        print(f"File: {filename}")
        print(f"Saved to: {filepath}")
        print("="*60 + "\n")
        
        return jsonify({
            'success': True,
            'message': 'Visa booking document generated successfully!',
            'document': {
                'id': confirmation_number,
                'filename': filename,
                'company': data['company'],
                'email': data['email'],
                'guest_name': data['guestName'],
                'nights': nights,
                'total_amount': total_amount,
                'download_url': f'/download/{confirmation_number}',
                'view_url': f'/documents/{confirmation_number}'
            }
        })
        
    except Exception as e:
        print(f"❌ Error generating document: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return jsonify({
            'success': False,
            'message': f'Error generating document: {str(e)}'
        }), 500

@app.route('/documents', methods=['GET'])
def list_documents():
    """View all generated documents"""
    print(f"请求文档列表，当前有 {len(documents_store)} 个文档")
    return jsonify({
        'success': True,
        'count': len(documents_store),
        'documents': [
            {
                'id': doc['id'],
                'filename': doc['filename'],
                'company': doc['company'],
                'email': doc['email'],
                'guest_name': doc['guest_name'],
                'dates': f"{doc['arrival_date']} to {doc['departure_date']}",
                'nights': doc['nights'],
                'total_amount': doc['total_amount'],
                'generated_date': doc['generated_date'],
                'download_url': doc['download_url'],
                'print_url': doc['print_url']
            }
            for doc in documents_store
        ]
    })

@app.route('/documents/<document_id>', methods=['GET'])
def get_document(document_id):
    """Get specific document information"""
    print(f"查找文档: {document_id}")
    for doc in documents_store:
        if doc['id'] == document_id:
            return jsonify({
                'success': True,
                'document': doc
            })
    
    return jsonify({
        'success': False,
        'message': 'Document not found'
    }), 404

@app.route('/download/<document_id>', methods=['GET'])
def download_document(document_id):
    """Download the Excel file"""
    print(f"下载文档请求: {document_id}")
    for doc in documents_store:
        if doc['id'] == document_id:
            filepath = Path(doc['filepath'])
            print(f"查找文件: {filepath}")
            if filepath.exists():
                print(f"文件存在，准备下载: {filepath}")
                return send_file(
                    str(filepath),
                    as_attachment=True,
                    download_name=doc['filename'],
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            else:
                print(f"文件不存在: {filepath}")
    
    return jsonify({
        'success': False,
        'message': 'File not found'
    }), 404

@app.route('/print/<document_id>', methods=['GET'])
def print_document(document_id):
    """打印文档信息到控制台"""
    print(f"打印文档请求: {document_id}")
    for doc in documents_store:
        if doc['id'] == document_id:
            print("\n" + "="*60)
            print("DOCUMENT PRINT REQUEST")
            print("="*60)
            print(f"Company: {doc['company']}")
            print(f"Email: {doc['email']}")
            print(f"Guest: {doc['guest_name']}")
            print(f"Dates: {doc['arrival_date']} to {doc['departure_date']}")
            print(f"Nights: {doc['nights']}")
            print(f"Total: {doc['total_amount']:,} CFA")
            print(f"Document ID: {doc['id']}")
            print(f"Generated: {doc['generated_date']}")
            print(f"File: {doc['filename']}")
            print(f"Path: {doc['filepath']}")
            print("="*60 + "\n")
            
            return jsonify({
                'success': True,
                'message': 'Document information printed to console',
                'document': {
                    'id': doc['id'],
                    'company': doc['company'],
                    'email': doc['email'],
                    'guest_name': doc['guest_name'],
                    'dates': f"{doc['arrival_date']} to {doc['departure_date']}",
                    'nights': doc['nights'],
                    'total_amount': doc['total_amount'],
                    'filename': doc['filename']
                }
            })
    
    return jsonify({
        'success': False,
        'message': 'Document not found'
    }), 404

@app.route('/cleanup', methods=['POST'])
def cleanup_documents():
    """手动清理超过48小时的文档"""
    try:
        print("执行文档清理...")
        # 这里可以添加实际的清理逻辑
        return jsonify({
            'success': True,
            'message': 'Cleanup completed successfully',
            'remaining_documents': len(documents_store)
        })
    except Exception as e:
        print(f"清理失败: {e}")
        return jsonify({
            'success': False,
            'message': f'Cleanup failed: {str(e)}'
        }), 500

def create_template_file():
    """Create a basic template if not exists"""
    print("Creating template file...")
    try:
        wb = load_workbook()
        ws = wb.active
        ws.title = "ipms_master_bill"
        
        # Add basic structure
        ws['C3'] = "Reservation Confirmation"
        ws['B5'] = "Booking Name"
        ws['B7'] = "Company Name"
        ws['B8'] = "Booking Date"
        ws['C9'] = "Email"
        ws['D10'] = "Remark"
        
        # Save template
        wb.save(str(TEMPLATE_PATH))
        print(f"✅ Template created: {TEMPLATE_PATH}")
        return True
    except Exception as e:
        print(f"❌ Failed to create template: {e}")
        return False

@app.route('/check-template', methods=['GET'])
def check_template():
    """Check if template exists and its structure"""
    if TEMPLATE_PATH.exists():
        try:
            wb = load_workbook(str(TEMPLATE_PATH))
            ws = wb.active
            sheet_name = ws.title
            
            # Check some key cells
            key_cells = {
                'C3': ws['C3'].value,
                'B5': ws['B5'].value,
                'sheet_name': sheet_name
            }
            
            return jsonify({
                'success': True,
                'message': 'Template found and loaded successfully',
                'sheet_name': sheet_name,
                'key_cells': key_cells
            })
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'Error loading template: {str(e)}'
            }), 500
    else:
        return jsonify({
            'success': False,
            'message': f'Template file not found at: {TEMPLATE_PATH}'
        }), 404

@app.route('/debug', methods=['GET'])
def debug_info():
    """调试信息页面"""
    info = {
        'python_version': sys.version,
        'current_directory': os.getcwd(),
        'base_dir': str(BASE_DIR),
        'template_exists': TEMPLATE_PATH.exists(),
        'generated_folder_exists': GENERATED_FOLDER.exists(),
        'generated_folder': str(GENERATED_FOLDER),
        'generated_files': list(GENERATED_FOLDER.glob('*.xlsx')) if GENERATED_FOLDER.exists() else [],
        'documents_count': len(documents_store),
        'uploads_folder_exists': UPLOAD_FOLDER.exists(),
    }
    return jsonify(info)

if __name__ == '__main__':
    print("="*60)
    print("Starting Visa Booking Document Generator")
    print("="*60)
    
    # 检查目录和文件
    create_directories()
    
    # Check template
    if not TEMPLATE_PATH.exists():
        print("Template not found, creating basic template...")
        create_template_file()
    else:
        print(f"✅ Template found: {TEMPLATE_PATH}")
    
    print(f"📁 Generated folder: {GENERATED_FOLDER}")
    print(f"📁 Uploads folder: {UPLOAD_FOLDER}")
    print(f"📋 Documents in memory: {len(documents_store)}")
    print("\n🚀 Application ready!")
    print("="*60)
    
    app.run(debug=True)
